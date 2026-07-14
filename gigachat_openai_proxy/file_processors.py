from __future__ import annotations

from io import BytesIO
from typing import Any

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


MAX_EXCEL_OPERATIONS = 2_000
MAX_EXCEL_FILE_BYTES = 25 * 1024 * 1024


class FileProcessorError(ValueError):
    pass


def edit_xlsx_bytes(
    content: bytes,
    operations_payload: Any,
) -> tuple[bytes, dict[str, Any]]:
    if len(content) > MAX_EXCEL_FILE_BYTES:
        raise FileProcessorError(
            f"xlsx file is too large; max size is {MAX_EXCEL_FILE_BYTES} bytes"
        )

    operations = normalize_operations(operations_payload)
    workbook = load_workbook(BytesIO(content))

    applied: list[str] = []
    for operation in operations:
        operation_type = operation.get("op")
        if operation_type == "create_sheet":
            create_sheet(workbook, operation)
        elif operation_type == "rename_sheet":
            rename_sheet(workbook, operation)
        elif operation_type == "delete_sheet":
            delete_sheet(workbook, operation)
        elif operation_type == "set_cell":
            set_cell(workbook, operation)
        elif operation_type == "append_row":
            append_row(workbook, operation)
        elif operation_type == "set_column_width":
            set_column_width(workbook, operation)
        elif operation_type == "style_cell":
            style_cell(workbook, operation)
        else:
            raise FileProcessorError(f"unsupported Excel operation: {operation_type}")
        applied.append(str(operation_type))

    output = BytesIO()
    workbook.save(output)
    return output.getvalue(), {
        "operation_count": len(operations),
        "operations": applied,
        "sheet_names": workbook.sheetnames,
    }


def normalize_operations(payload: Any) -> list[dict[str, Any]]:
    if isinstance(payload, dict):
        operations = payload.get("operations")
    else:
        operations = payload

    if not isinstance(operations, list):
        raise FileProcessorError("operations must be a JSON array")
    if len(operations) > MAX_EXCEL_OPERATIONS:
        raise FileProcessorError(
            f"too many operations; max is {MAX_EXCEL_OPERATIONS}")

    normalized: list[dict[str, Any]] = []
    for index, operation in enumerate(operations):
        if not isinstance(operation, dict):
            raise FileProcessorError(f"operation #{index} must be an object")
        if not isinstance(operation.get("op"), str):
            raise FileProcessorError(f"operation #{index} must contain string op")
        normalized.append(operation)
    return normalized


def worksheet(workbook: Any, operation: dict[str, Any]) -> Any:
    sheet_name = operation.get("sheet")
    if sheet_name is None:
        return workbook.active
    if not isinstance(sheet_name, str) or not sheet_name:
        raise FileProcessorError("sheet must be a non-empty string")
    if sheet_name not in workbook.sheetnames:
        raise FileProcessorError(f"sheet does not exist: {sheet_name}")
    return workbook[sheet_name]


def required_string(operation: dict[str, Any], key: str) -> str:
    value = operation.get(key)
    if not isinstance(value, str) or not value:
        raise FileProcessorError(f"{key} must be a non-empty string")
    return value


def create_sheet(workbook: Any, operation: dict[str, Any]) -> None:
    title = required_string(operation, "title")
    if title in workbook.sheetnames:
        raise FileProcessorError(f"sheet already exists: {title}")
    workbook.create_sheet(title=title)


def rename_sheet(workbook: Any, operation: dict[str, Any]) -> None:
    ws = worksheet(workbook, operation)
    title = required_string(operation, "title")
    if title in workbook.sheetnames:
        raise FileProcessorError(f"sheet already exists: {title}")
    ws.title = title


def delete_sheet(workbook: Any, operation: dict[str, Any]) -> None:
    ws = worksheet(workbook, operation)
    if len(workbook.worksheets) <= 1:
        raise FileProcessorError("cannot delete the last sheet")
    workbook.remove(ws)


def set_cell(workbook: Any, operation: dict[str, Any]) -> None:
    ws = worksheet(workbook, operation)
    cell = required_string(operation, "cell")
    ws[cell].value = operation.get("value")


def append_row(workbook: Any, operation: dict[str, Any]) -> None:
    ws = worksheet(workbook, operation)
    values = operation.get("values")
    if not isinstance(values, list):
        raise FileProcessorError("values must be an array")
    ws.append(values)


def set_column_width(workbook: Any, operation: dict[str, Any]) -> None:
    ws = worksheet(workbook, operation)
    column = operation.get("column")
    if isinstance(column, int):
        column = get_column_letter(column)
    if not isinstance(column, str) or not column:
        raise FileProcessorError("column must be a column letter or number")

    width = operation.get("width")
    if not isinstance(width, int | float) or width <= 0:
        raise FileProcessorError("width must be a positive number")
    ws.column_dimensions[column.upper()].width = float(width)


def style_cell(workbook: Any, operation: dict[str, Any]) -> None:
    ws = worksheet(workbook, operation)
    cell = ws[required_string(operation, "cell")]

    font_kwargs: dict[str, Any] = {}
    for key in ("bold", "italic", "underline"):
        if isinstance(operation.get(key), bool):
            font_kwargs[key] = operation[key]
    font_color = operation.get("font_color")
    if isinstance(font_color, str) and font_color:
        font_kwargs["color"] = normalized_color(font_color)
    if font_kwargs:
        cell.font = Font(
            name=cell.font.name,
            size=cell.font.size,
            bold=font_kwargs.get("bold", cell.font.bold),
            italic=font_kwargs.get("italic", cell.font.italic),
            underline=font_kwargs.get("underline", cell.font.underline),
            color=font_kwargs.get("color", cell.font.color),
        )

    fill_color = operation.get("fill_color")
    if isinstance(fill_color, str) and fill_color:
        cell.fill = PatternFill(
            fill_type="solid",
            fgColor=normalized_color(fill_color),
        )

    number_format = operation.get("number_format")
    if isinstance(number_format, str) and number_format:
        cell.number_format = number_format


def normalized_color(value: str) -> str:
    color = value.removeprefix("#").upper()
    if len(color) == 6:
        return f"FF{color}"
    if len(color) == 8:
        return color
    raise FileProcessorError("colors must be #RRGGBB or AARRGGBB")
